import os
import io
from datetime import datetime, timedelta
from flask import (Flask, render_template, request, redirect, url_for,
                   flash, session, send_file, jsonify)
from flask_sqlalchemy import SQLAlchemy
from flask_bcrypt import Bcrypt
from openpyxl import Workbook
from io import BytesIO
from openpyxl.styles import PatternFill
from functools import wraps
import jwt
from flask_cors import CORS

# --- APP CONFIGURATION ---
app = Flask(__name__)
CORS(app)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'a_very_secure_and_random_secret_key_for_dev')

app.config['JWT_SECRET_KEY'] = os.environ.get('JWT_SECRET_KEY','a-different-super-strong-secret-for-dev')

# Use PostgreSQL if DATABASE_URL is set (in production), otherwise use SQLite (for local development)
DATABASE_URL = os.environ.get('DATABASE_URL')
if DATABASE_URL:
    # The Neon URL starts with postgres://, but SQLAlchemy needs postgresql://
    app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL.replace("postgres://", "postgresql://", 1)
else:
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///stratabet.db'

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)
bcrypt = Bcrypt(app)


# --- DATABASE MODELS (Unchanged)---
class User(db.Model):
    roll_number = db.Column(db.String(20), primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    password_hash = db.Column(db.String(60), nullable=False)
    points = db.Column(db.Integer, nullable=False, default=200)
    is_admin = db.Column(db.Boolean, nullable=False, default=False)
    bets = db.relationship('Bet', backref='bettor', lazy=True, cascade="all, delete-orphan")

    @property
    def password(self):
        raise AttributeError('password is not a readable attribute')

    @password.setter
    def password(self, password):
        self.password_hash = bcrypt.generate_password_hash(password).decode('utf-8')

    def verify_password(self, password):
        return bcrypt.check_password_hash(self.password_hash, password)
        
    def to_dict(self):
        return {
            'roll_number': self.roll_number,
            'name': self.name,
            'points': self.points,
            'is_admin': self.is_admin
        }


class Event(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False, unique=True)
    is_active = db.Column(db.Boolean, nullable=False, default=True)
    questions = db.relationship('Question', backref='event', lazy=True, cascade="all, delete-orphan")

    # NEW: Method to serialize object to a dictionary
    def to_dict(self):
        return {
            'id': self.id,
            'name': self.name,
            'is_active': self.is_active,
            'questions': [q.to_dict() for q in self.questions]
        }


class Question(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    event_id = db.Column(db.Integer, db.ForeignKey('event.id'), nullable=False)
    text = db.Column(db.String(255), nullable=False)
    is_open = db.Column(db.Boolean, nullable=False, default=True)
    winning_option_id = db.Column(db.Integer, db.ForeignKey('option.id'), nullable=True)
    options = db.relationship('Option', foreign_keys='Option.question_id', backref='question', lazy=True, cascade="all, delete-orphan")
    winning_option = db.relationship('Option', foreign_keys=[winning_option_id])
    bets = db.relationship('Bet', backref='question', lazy=True, cascade="all, delete-orphan")

    # NEW: Method to serialize object to a dictionary
    def to_dict(self):
        return {
            'id': self.id,
            'event_id': self.event_id,
            'event_name': self.event.name,
            'text': self.text,
            'is_open': self.is_open,
            'winning_option_id': self.winning_option_id,
            'options': [opt.to_dict() for opt in self.options]
        }


class Option(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    question_id = db.Column(db.Integer, db.ForeignKey('question.id'), nullable=False)
    text = db.Column(db.String(100), nullable=False)
    odds = db.Column(db.Float, nullable=False, default=1.8)

    # NEW: Method to serialize object to a dictionary
    def to_dict(self):
        return {
            'id': self.id,
            'question_id': self.question_id,
            'text': self.text,
            'odds': self.odds
        }


class Bet(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_roll_number = db.Column(db.String(20), db.ForeignKey('user.roll_number'), nullable=False)
    question_id = db.Column(db.Integer, db.ForeignKey('question.id'), nullable=False)
    option_id = db.Column(db.Integer, db.ForeignKey('option.id'), nullable=False)
    amount = db.Column(db.Integer, nullable=False)
    status = db.Column(db.String(20), nullable=False, default='Pending')
    timestamp = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    option = db.relationship('Option', backref='bets')

    # NEW: Method to serialize object to a dictionary
    def to_dict(self):
        return {
            'id': self.id,
            'user_roll_number': self.user_roll_number,
            'question_id': self.question_id,
            'question_text': self.question.text,  # Added for convenience
            'option_id': self.option_id,
            'option_text': self.option.text,      # Added for convenience
            'amount': self.amount,
            'status': self.status,
            'timestamp': self.timestamp.isoformat() # Use ISO format for dates
        }


class Team(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), unique=True, nullable=False)
    squad = db.Column(db.Text, nullable=True)

    # NEW: Method to serialize object to a dictionary
    def to_dict(self):
        return {
            'id': self.id,
            'name': self.name,
            'squad': self.squad
        }
# --- HELPER FUNCTIONS & SETUP (Unchanged) ---
def get_current_user():
    if 'roll_number' in session:
        return User.query.get(session['roll_number'])
    return None

def admin_required(f):
    def decorated_function(*args, **kwargs):
        user = get_current_user()
        if not user or not user.is_admin:
            flash("You do not have permission to access this page.", "danger")
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

# NEW: Token decorator for API
def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = None
        if 'Authorization' in request.headers:
            # Expected format: "Bearer <token>"
            token = request.headers['Authorization'].split(" ")[1]
        if not token:
            return jsonify({'message': 'Token is missing!'}), 401
        try:
            data = jwt.decode(token, app.config['JWT_SECRET_KEY'], algorithms=["HS256"])
            current_user = User.query.get(data['roll_number'])
            if not current_user:
                return jsonify({'message': 'User not found!'}), 401
        except jwt.ExpiredSignatureError:
            return jsonify({'message': 'Token has expired!'}), 401
        except jwt.InvalidTokenError:
            return jsonify({'message': 'Token is invalid!'}), 401
        
        return f(current_user, *args, **kwargs)
    return decorated

def get_ranked_leaderboard():
    """Queries players and calculates their rank, handling ties."""
    players = User.query.filter(User.bets.any(), User.is_admin == False).order_by(db.desc(User.points)).all()
    
    ranked_players = []
    last_score = -1
    last_rank = 0
    
    for i, player in enumerate(players):
        current_rank = i + 1
        if player.points == last_score:
            # Same score as the previous player, use their rank
            ranked_players.append({'rank': last_rank, 'player': player})
        else:
            # New score, rank is the current position (1-based index)
            last_rank = current_rank
            last_score = player.points
            ranked_players.append({'rank': last_rank, 'player': player})
            
    return ranked_players

@app.cli.command("init-db")
def init_db_command():
    with app.app_context():
        db.create_all()
        if not User.query.filter_by(roll_number='admin').first():
            admin_user = User(roll_number='admin', name='Admin User', is_admin=True)
            admin_user.password = 'password'
            db.session.add(admin_user)
            db.session.commit()
            print("Database initialized and admin user 'admin' created.")
        
        teams = ['Alpha Wolves', 'Black Pirates', 'Dragon Slayers', 'Viking Warriors']
        for team_name in teams:
            if not Team.query.filter_by(name=team_name).first():
                team = Team(name=team_name, squad="Player 1, Player 2, Player 3...")
                db.session.add(team)
        db.session.commit()
        print("Teams have been populated.")


# --- CORE & USER ROUTES ---
@app.route('/')
def index():
    if get_current_user():
        return redirect(url_for('dashboard'))
    return render_template('index.html')

@app.route('/leaderboard')
def leaderboard():
    user = get_current_user()
    ranked_players = get_ranked_leaderboard()
    return render_template('leaderboard.html', user=user, players=ranked_players)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        roll_number = request.form['roll_number']
        password = request.form['password']
        roll_number = roll_number.replace('/', '')
        user = User.query.get(roll_number)

        if user and user.verify_password(password):
            session['roll_number'] = user.roll_number
            flash('Login successful!', 'success')
            if user.is_admin:
                return redirect(url_for('admin_dashboard'))
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid roll number or password.', 'danger')
    return render_template('login.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        name = request.form['name']
        roll_number = request.form['roll_number']
        password = request.form['password']
        roll_number = roll_number.replace('/', '')
        
        if User.query.get(roll_number):
            flash('This roll number is already registered.', 'warning')
            return redirect(url_for('login'))

        new_user = User(name=name, roll_number=roll_number)
        new_user.password = password
        db.session.add(new_user)
        db.session.commit()
        
        flash('Registration successful! Please log in.', 'success')
        return redirect(url_for('login'))
        
    return render_template('register.html')

@app.route('/logout')
def logout():
    session.pop('roll_number', None)
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))

@app.route('/change-password', methods=['GET', 'POST'])
def change_password():
    user = get_current_user()
    if not user:
        return redirect(url_for('login'))

    if request.method == 'POST':
        current_password = request.form['current_password']
        new_password = request.form['new_password']
        confirm_password = request.form['confirm_password']

        if not user.verify_password(current_password):
            flash('Your current password is incorrect.', 'danger')
            return redirect(url_for('change_password'))
        
        if new_password != confirm_password:
            flash('New passwords do not match.', 'danger')
            return redirect(url_for('change_password'))
        
        user.password = new_password
        db.session.commit()
        flash('Your password has been updated successfully!', 'success')
        return redirect(url_for('dashboard'))

    return render_template('change_password.html', user=user)

@app.route('/dashboard')
def dashboard():
    user = get_current_user()
    if not user:
        return redirect(url_for('login'))
    
    active_events = Event.query.filter_by(is_active=True).all()
    user_bets_q_ids = [bet.question_id for bet in user.bets]
    
    available_questions = {}
    for event in active_events:
        questions = Question.query.filter(
            Question.event_id == event.id,
            Question.is_open == True,
            Question.id.notin_(user_bets_q_ids)
        ).all()
        if questions:
            available_questions[event] = questions

    return render_template('dashboard.html', user=user, available_questions=available_questions)

@app.route('/place_bet/<int:question_id>', methods=['POST'])
def place_bet(question_id):
    user = get_current_user()
    if not user:
        return redirect(url_for('login'))
    question = Question.query.get_or_404(question_id)
    if not question.is_open:
        flash("Betting for this question is now closed.", "warning")
        return redirect(url_for('dashboard'))

    try:
        amount = int(request.form['amount'])
        option_id = int(request.form['option_id'])
    except (ValueError, KeyError):
        flash("Invalid bet data submitted.", "danger")
        return redirect(url_for('dashboard'))

    if amount <= 0:
        flash("Bet amount must be positive.", "warning")
        return redirect(url_for('dashboard'))

    if user.points < amount:
        flash("You do not have enough points for this bet.", "danger")
        return redirect(url_for('dashboard'))

    existing_bet = Bet.query.filter_by(user_roll_number=user.roll_number, question_id=question_id).first()
    if existing_bet:
        flash("You have already placed a bet on this question.", "warning")
        return redirect(url_for('dashboard'))

    user.points -= amount
    new_bet = Bet(user_roll_number=user.roll_number, question_id=question_id, option_id=option_id, amount=amount)
    db.session.add(new_bet)
    db.session.commit()

    flash(f"Bet of {amount} points placed successfully!", "success")
    return redirect(url_for('dashboard'))

@app.route('/my_bets')
def my_bets():
    user = get_current_user()
    if not user:
        return redirect(url_for('login'))
    
    bets = Bet.query.filter_by(user_roll_number=user.roll_number).order_by(Bet.timestamp.desc()).all()
    return render_template('my_bets.html', user=user, bets=bets)

@app.route('/squads')
def squads():
    user = get_current_user()
    if not user:
        return redirect(url_for('login'))
    teams = Team.query.all()
    return render_template('squads.html', teams=teams, user=user)

# --- ADMIN PANEL ---
@app.route('/admin')
@admin_required
def admin_dashboard():
    user = get_current_user()
    all_events = Event.query.order_by(Event.name).all()
    return render_template('admin/dashboard.html', user=user, events=all_events)

@app.route('/admin/reset_user', methods=['POST'])
@admin_required
def reset_user():
    roll_number = request.form.get('roll_number_to_reset')
    user_to_reset = User.query.get(roll_number)

    if not user_to_reset:
        flash(f"User with roll number '{roll_number}' not found.", 'danger')
        return redirect(url_for('admin_dashboard'))
    
    user_to_reset.password = 'password'
    
    db.session.commit()

    flash(f"Account for {user_to_reset.name} ({user_to_reset.roll_number}) has been reset. New password is 'password'.", 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/events/create', methods=['POST'])
@admin_required
def create_event():
    name = request.form.get('name')
    if name and not Event.query.filter_by(name=name).first():
        new_event = Event(name=name)
        db.session.add(new_event)
        db.session.commit()
        flash(f'Event "{name}" created successfully.', 'success')
    else:
        flash('Event name is required and must be unique.', 'danger')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/events/delete/<int:event_id>', methods=['POST'])
@admin_required
def delete_event(event_id):
    event_to_delete = Event.query.get_or_404(event_id)
    db.session.delete(event_to_delete)
    db.session.commit()
    flash(f'Event "{event_to_delete.name}" and all its data have been permanently deleted.', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/events/toggle/<int:event_id>')
@admin_required
def toggle_event_status(event_id):
    event = Event.query.get_or_404(event_id)
    event.is_active = not event.is_active
    db.session.commit()
    status = "activated" if event.is_active else "deactivated"
    flash(f'Event "{event.name}" has been {status}.', 'info')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/questions/<int:event_id>')
@admin_required
def manage_questions(event_id):
    user = get_current_user()
    event = Event.query.get_or_404(event_id)
    return render_template('admin/questions.html', user=user, event=event)
    
@app.route('/admin/questions/create/<int:event_id>', methods=['POST'])
@admin_required
def create_question(event_id):
    event = Event.query.get_or_404(event_id)
    question_text = request.form.get('question_text')

    if not question_text:
        flash('Question text cannot be empty.', 'danger')
        return redirect(url_for('manage_questions', event_id=event.id))

    option_texts = request.form.getlist('option_text')
    option_odds_list = request.form.getlist('option_odds')

    valid_options_provided = any(text and odds for text, odds in zip(option_texts, option_odds_list))
    if not valid_options_provided:
        flash('You must provide at least one complete option (text and odds).', 'danger')
        return redirect(url_for('manage_questions', event_id=event.id))

    new_question = Question(text=question_text, event_id=event.id)
    db.session.add(new_question)
    
    options_added_count = 0
    for text, odds_str in zip(option_texts, option_odds_list):
        if text and odds_str:
            try:
                odds = float(odds_str)
                if odds <= 1.0:
                    flash(f'Odds for "{text}" must be greater than 1.0. This option was not added.', 'warning')
                    continue
                
                option = Option(text=text, odds=odds, question=new_question)
                db.session.add(option)
                options_added_count += 1
            except ValueError:
                flash(f'Invalid odds value "{odds_str}". This option was not added.', 'warning')
    
    if options_added_count < 2:
        db.session.rollback() 
        flash('A question requires at least two valid options. The question was not created.', 'danger')
        return redirect(url_for('manage_questions', event_id=event.id))

    db.session.commit()
    flash(f'New question with {options_added_count} option(s) has been added.', 'success')
    return redirect(url_for('manage_questions', event_id=event.id))

# ** NEW: Route to delete a question **
@app.route('/admin/questions/delete/<int:question_id>', methods=['POST'])
@admin_required
def delete_question(question_id):
    question = Question.query.get_or_404(question_id)
    event_id = question.event_id
    question_text = question.text

    db.session.delete(question)
    db.session.commit()

    flash(f'Question "{question_text[:30]}..." and all associated data have been permanently deleted.', 'success')
    return redirect(url_for('manage_questions', event_id=event_id))


@app.route('/admin/questions/toggle/<int:question_id>')
@admin_required
def toggle_question_status(question_id):
    question = Question.query.get_or_404(question_id)
    question.is_open = not question.is_open
    db.session.commit()
    status = "opened for betting" if question.is_open else "closed for betting"
    flash(f'Question "{question.text[:30]}..." has been {status}.', 'info')
    return redirect(url_for('manage_questions', event_id=question.event_id))

@app.route('/admin/squads', methods=['GET', 'POST'])
@admin_required
def manage_squads():
    user = get_current_user()
    if request.method == 'POST':
        team_id = request.form.get('team_id')
        squad_text = request.form.get('squad_text')
        team = Team.query.get(team_id)
        if team:
            team.squad = squad_text
            db.session.commit()
            flash(f"Squad for {team.name} updated.", "success")
        return redirect(url_for('manage_squads'))

    teams = Team.query.all()
    return render_template('admin/squads.html', user=user, teams=teams)

@app.route('/admin/results')
@admin_required
def manage_results():
    user = get_current_user()
    unresolved_questions = Question.query.filter_by(is_open=False, winning_option_id=None).all()
    return render_template('admin/results.html', user=user, questions=unresolved_questions)

@app.route('/admin/results/process/<int:question_id>', methods=['POST'])
@admin_required
def process_results(question_id):
    question = Question.query.get_or_404(question_id)
    winning_option_id = request.form.get('winning_option_id')

    if not winning_option_id:
        flash('You must select a winning option.', 'danger')
        return redirect(url_for('manage_results'))

    question.winning_option_id = int(winning_option_id)
    winning_option = Option.query.get(question.winning_option_id)

    bets_to_process = Bet.query.filter_by(question_id=question_id, status='Pending').all()
    
    for bet in bets_to_process:
        if bet.option_id == question.winning_option_id:
            winnings = bet.amount * winning_option.odds
            bettor = User.query.get(bet.user_roll_number)
            bettor.points += int(winnings)
            bet.status = 'Won'
        else:
            bet.status = 'Lost'
    
    db.session.commit()
    flash(f'Results for question "{question.text[:30]}..." processed. {len(bets_to_process)} bets updated.', 'success')
    return redirect(url_for('manage_results'))


# --- ADMIN DOWNLOAD ROUTES (Unchanged) ---
@app.route('/admin/download_bets')
@admin_required
def download_bets():
    wb = Workbook()
    wb.remove(wb.active)
    events = Event.query.order_by(Event.id).all()
    color_palette = [
        PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
        PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        PatternFill(start_color="EAD1DC", end_color="EAD1DC", fill_type="solid")
    ]
    static_header_fill = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
    summary_header_fill = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")
    for event in events:
        ws = wb.create_sheet(title=event.name)
        questions = sorted(event.questions, key=lambda q: q.id)
        headers = ["Timestamp", "Roll Number", "Name"]
        for q in questions:
            headers.extend([
                f"Q{q.id}: {q.text}", "Selected Option", "Correct Answer", 
                "Bet Amount", "Odds"
            ])
        headers.extend(["Total Won/Lost in Event", "Final Score"])
        ws.append(headers)
        header_row = ws[1]
        for i in range(1, 4):
            header_row[i-1].fill = static_header_fill
        current_col = 4
        for i, q in enumerate(questions):
            color = color_palette[i % len(color_palette)]
            for _ in range(5):
                header_row[current_col-1].fill = color
                current_col += 1
        header_row[current_col-1].fill = summary_header_fill
        header_row[current_col].fill = summary_header_fill
        bets_in_event = Bet.query.join(Question).filter(Question.event_id == event.id).all()
        user_event_data = {}
        for bet in bets_in_event:
            roll_number = bet.user_roll_number
            if roll_number not in user_event_data:
                user_event_data[roll_number] = {
                    'user': bet.bettor, 'timestamp': bet.timestamp.strftime("%Y-%m-%d %H:%M"),
                    'bets': {}
                }
            user_event_data[roll_number]['bets'][bet.question_id] = bet
        for roll_number, data in user_event_data.items():
            user = data['user']
            total_won_lost = 0
            row_data = [data['timestamp'], user.roll_number, user.name]
            for q in questions:
                bet = data['bets'].get(q.id)
                if bet:
                    if bet.status == 'Won':
                        total_won_lost += (bet.amount * bet.option.odds) - bet.amount
                    elif bet.status == 'Lost':
                        total_won_lost -= bet.amount
                    row_data.extend([
                        "", bet.option.text,
                        bet.question.winning_option.text if bet.question.winning_option else "Pending",
                        bet.amount, bet.option.odds
                    ])
                else:
                    row_data.extend(["", "No Bet", "N/A", "", ""])
            row_data.append(int(total_won_lost))
            row_data.append(user.points)
            ws.append(row_data)
    excel_io = BytesIO()
    wb.save(excel_io)
    excel_io.seek(0)
    return send_file(
        excel_io,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'stratabet_bets_export_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )

@app.route('/admin/download_results')
@admin_required
def download_results():
    ranked_players = get_ranked_leaderboard()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Leaderboard"
    headers = ["Rank", "Name", "Roll Number", "Points"]
    ws.append(headers)

    header_fill = PatternFill(start_color="FFBF00", end_color="FFBF00", fill_type="solid")
    for cell in ws[1]:
        cell.fill = header_fill

    for item in ranked_players:
        ws.append([item['rank'], item['player'].name, item['player'].roll_number, item['player'].points])
    
    excel_io = BytesIO()
    wb.save(excel_io)
    excel_io.seek(0)
    
    return send_file(
        excel_io,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'stratabet_leaderboard_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )
#! Android APIs

@app.route('/api/register', methods=['POST'])
def api_register():
    data = request.get_json()
    if not data or not all(k in data for k in ('name', 'roll_number', 'password')):
        return jsonify({'message': 'Missing data'}), 400

    roll_number = data['roll_number'].replace('/', '')
    if User.query.get(roll_number):
        return jsonify({'message': 'This roll number is already registered'}), 409

    new_user = User(name=data['name'], roll_number=roll_number)
    new_user.password = data['password']
    db.session.add(new_user)
    db.session.commit()

    return jsonify({'message': 'Registration successful! Please log in.'}), 201


@app.route('/api/login', methods=['POST'])
def api_login():
    data = request.get_json()
    if not data or not all(k in data for k in ('roll_number', 'password')):
        return make_response('Could not verify', 401, {'WWW-Authenticate': 'Basic realm="Login required!"'})

    roll_number = data['roll_number'].replace('/', '')
    user = User.query.get(roll_number)

    if not user or not user.verify_password(data['password']):
        return jsonify({'message': 'Invalid roll number or password'}), 401

    token = jwt.encode({
        'roll_number': user.roll_number,
        'exp': datetime.utcnow() + timedelta(days=30) # Token valid for 30 days
    }, app.config['JWT_SECRET_KEY'], algorithm="HS256")

    return jsonify({'token': token, 'user': user.to_dict()})


@app.route('/api/dashboard', methods=['GET'])
@token_required
def api_dashboard(current_user):
    active_events = Event.query.filter_by(is_active=True).all()
    user_bets_q_ids = [bet.question_id for bet in current_user.bets]
    
    available_questions_list = []
    for event in active_events:
        questions = Question.query.filter(
            Question.event_id == event.id,
            Question.is_open == True,
            Question.id.notin_(user_bets_q_ids)
        ).all()
        if questions:
            for q in questions:
                 available_questions_list.append(q.to_dict())

    return jsonify(available_questions_list)

@app.route('/api/bets/place/<int:question_id>', methods=['POST'])
@token_required
def api_place_bet(current_user, question_id):
    data = request.get_json()
    if not data or not all(k in data for k in ('amount', 'option_id')):
        return jsonify({'message': 'Missing amount or option_id'}), 400

    question = Question.query.get_or_404(question_id)
    if not question.is_open:
        return jsonify({'message': 'Betting for this question is now closed'}), 403

    try:
        amount = int(data['amount'])
        option_id = int(data['option_id'])
    except ValueError:
        return jsonify({'message': 'Invalid bet data submitted'}), 400

    if amount <= 0:
        return jsonify({'message': 'Bet amount must be positive'}), 400

    if current_user.points < amount:
        return jsonify({'message': 'You do not have enough points for this bet'}), 402

    if Bet.query.filter_by(user_roll_number=current_user.roll_number, question_id=question_id).first():
        return jsonify({'message': 'You have already placed a bet on this question'}), 409

    current_user.points -= amount
    new_bet = Bet(user_roll_number=current_user.roll_number, question_id=question_id, option_id=option_id, amount=amount)
    db.session.add(new_bet)
    db.session.commit()

    return jsonify({'message': f'Bet of {amount} points placed successfully!', 'new_points': current_user.points}), 201


@app.route('/api/my-bets', methods=['GET'])
@token_required
def api_my_bets(current_user):
    bets = Bet.query.filter_by(user_roll_number=current_user.roll_number).order_by(Bet.timestamp.desc()).all()
    return jsonify([bet.to_dict() for bet in bets])


@app.route('/api/leaderboard', methods=['GET'])
def api_leaderboard():
    ranked_players = get_ranked_leaderboard()
    
    # Serialize the data into the desired JSON format
    leaderboard_data = []
    for item in ranked_players:
        leaderboard_data.append({
            'rank': item['rank'],
            'user': item['player'].to_dict()
        })
        
    return jsonify(leaderboard_data)

@app.route('/api/squads', methods=['GET'])
@token_required
def api_squads(current_user):
    teams = Team.query.all()
    return jsonify([team.to_dict() for team in teams])


if __name__ == '__main__':
    app.run(debug=True)